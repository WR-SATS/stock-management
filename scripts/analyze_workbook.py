from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


if len(sys.argv) < 2:
    raise SystemExit("Usage: python scripts/analyze_workbook.py <path-to-workbook.xlsx>")

INPUT_PATH = Path(sys.argv[1])
OUTPUT_DIR = Path("docs")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def clean(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, (int, float, bool)):
        return value
    return str(value).strip() if isinstance(value, str) else str(value)


def row_values(ws, row_number, max_col):
    return [clean(ws.cell(row_number, col).value) for col in range(1, max_col + 1)]


def compact_rows(rows):
    compacted = []
    for row in rows:
        if any(value not in (None, "") for value in row):
            compacted.append(row)
    return compacted


def infer_header_row(ws, scan_rows=15):
    best_row = 1
    best_score = -1
    limit = min(ws.max_row, scan_rows)
    for row_number in range(1, limit + 1):
        values = row_values(ws, row_number, min(ws.max_column, 60))
        non_empty = [v for v in values if v not in (None, "")]
        text_count = sum(isinstance(v, str) for v in non_empty)
        score = len(non_empty) + text_count * 0.5
        if score > best_score:
            best_score = score
            best_row = row_number
    return best_row


def profile_sheet(ws):
    max_col = min(ws.max_column, 80)
    header_row = infer_header_row(ws)
    header = row_values(ws, header_row, max_col)
    rows = compact_rows([row_values(ws, r, max_col) for r in range(1, min(ws.max_row, 25) + 1)])
    sample_after_header = compact_rows(
        [row_values(ws, r, max_col) for r in range(header_row + 1, min(ws.max_row, header_row + 11) + 1)]
    )

    column_samples = {}
    for col_index, title in enumerate(header, start=1):
        if title in (None, ""):
            continue
        values = []
        for row_number in range(header_row + 1, min(ws.max_row, header_row + 101) + 1):
            value = clean(ws.cell(row_number, col_index).value)
            if value not in (None, ""):
                values.append(value)
            if len(values) >= 8:
                break
        column_samples[str(title)] = values

    merged_ranges = [str(rng) for rng in list(ws.merged_cells.ranges)[:20]]

    formulas = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), max_col=max_col):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append({"cell": cell.coordinate, "formula": cell.value[:160]})
                if len(formulas) >= 20:
                    break
        if len(formulas) >= 20:
            break

    return {
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "inferred_header_row": header_row,
        "header": header,
        "top_rows": rows,
        "sample_after_header": sample_after_header,
        "column_samples": column_samples,
        "merged_ranges": merged_ranges,
        "formula_samples": formulas,
    }


def main():
    workbook = load_workbook(INPUT_PATH, data_only=False, read_only=False)
    profiles = [profile_sheet(ws) for ws in workbook.worksheets]
    (OUTPUT_DIR / "workbook-profile.json").write_text(
        json.dumps(profiles, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(
        [
            {
                "sheet": item["sheet"],
                "max_row": item["max_row"],
                "max_column": item["max_column"],
                "inferred_header_row": item["inferred_header_row"],
                "header": item["header"][:20],
            }
            for item in profiles
        ],
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()
